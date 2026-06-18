#!/usr/bin/env python3
"""Fetch OSWorld-Verified leaderboard data from the official Excel file."""

from __future__ import annotations

import argparse
import io
import json
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

OSWORLD_XLSX_URL = "https://os-world.github.io/static/data/osworld_verified_results.xlsx"
SHEET_NAME = "Eval Results"

# Foundation E2E GUI: no extra a11y tree, no extra coding actions, no multiple rollout, 100 steps
FOUNDATION_MAX_STEPS = 100
FOUNDATION_A11Y = "No"
FOUNDATION_CODING = "No"
FOUNDATION_ROLLOUT = "No"


def _excel_serial_to_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            # Excel epoch: 1899-12-30 (accounts for the Lotus 1-2-3 leap-year bug)
            d = (datetime(1899, 12, 30) + timedelta(days=int(value))).date()
            return d.isoformat()
        except (ValueError, OverflowError):
            return None
    return None


def fetch_xlsx_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def parse_rows(data: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet {SHEET_NAME!r} not found; available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    raw_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not raw_rows:
        return []

    headers = [str(h).strip() if h is not None else None for h in raw_rows[0]]

    def col(row: tuple[Any, ...], name: str) -> Any:
        try:
            return row[headers.index(name)]
        except ValueError:
            return None

    result = []
    for row in raw_rows[1:]:
        model = col(row, "Model")
        if not isinstance(model, str) or not model.strip():
            continue
        result.append(
            {
                "model": model.strip(),
                "institution": col(row, "Institution"),
                "approach_type": col(row, "Approach type"),
                "max_steps": col(row, "Max steps"),
                "a11y": col(row, "Additional a11y tree used"),
                "coding": col(row, "Additional coding-based action"),
                "rollout": col(row, "Multiple rollout"),
                "date": col(row, "Date"),
                "score": col(row, "Success rate"),
            }
        )
    return result


def is_foundation_e2e(row: dict[str, Any]) -> bool:
    return (
        row.get("max_steps") == FOUNDATION_MAX_STEPS
        and row.get("a11y") == FOUNDATION_A11Y
        and row.get("coding") == FOUNDATION_CODING
        and row.get("rollout") == FOUNDATION_ROLLOUT
    )


def aggregate(rows: list[dict[str, Any]], foundation_only: bool) -> list[dict[str, Any]]:
    filtered = [r for r in rows if not foundation_only or is_foundation_e2e(r)]
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in filtered:
        if isinstance(row["score"], (int, float)):
            groups[row["model"]].append(row)

    aggregated = []
    for model, model_rows in groups.items():
        scores = [r["score"] for r in model_rows]
        avg = round(sum(scores) / len(scores), 1)
        latest = max(model_rows, key=lambda r: _excel_serial_to_iso(r["date"]) or "")
        aggregated.append(
            {
                "model": model,
                "success_rate": avg,
                "date": _excel_serial_to_iso(latest["date"]),
                "approach_type": latest["approach_type"],
                "runs": len(scores),
            }
        )

    aggregated.sort(key=lambda r: r["success_rate"], reverse=True)
    return aggregated


def get_scores(foundation_only: bool = True) -> list[dict[str, Any]]:
    data = fetch_xlsx_bytes(OSWORLD_XLSX_URL)
    rows = parse_rows(data)
    return aggregate(rows, foundation_only=foundation_only)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch OSWorld-Verified leaderboard scores.")
    parser.add_argument(
        "--format",
        choices=["table", "json", "names"],
        default="table",
        help="Output format (default: table).",
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--foundation-only",
        action="store_true",
        default=True,
        help="Only include Foundation E2E GUI entries (default).",
    )
    group.add_argument(
        "--all",
        action="store_true",
        dest="all_entries",
        help="Include all entries regardless of setup.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    foundation_only = not args.all_entries
    scores = get_scores(foundation_only=foundation_only)

    if args.format == "json":
        print(json.dumps(scores, ensure_ascii=False))
    elif args.format == "names":
        for entry in scores:
            print(entry["model"])
    else:
        col_widths = [
            max(len("MODEL"), max((len(e["model"]) for e in scores), default=0)),
            7,
            10,
            4,
        ]
        fmt = f"{{:<{col_widths[0]}}}  {{:>{col_widths[1]}}}  {{:<{col_widths[2]}}}  {{}}"
        print(fmt.format("MODEL", "SCORE", "DATE", "APPROACH"))
        for entry in scores:
            print(
                fmt.format(
                    entry["model"],
                    str(entry["success_rate"]),
                    entry.get("date") or "",
                    entry.get("approach_type") or "",
                )
            )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print()
        raise SystemExit(130)
